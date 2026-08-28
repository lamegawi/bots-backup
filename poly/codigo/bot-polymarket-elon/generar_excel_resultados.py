#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GENERAR EXCEL DE RESULTADOS — a partir de un CSV de operaciones
================================================================
Lee un CSV de operaciones (formato de resultados_simulacion.csv) y
genera Resultados_Simulacion.xlsx con:

  · Resultados  — cada operación simulada (fecha, bin, lado, precio,
    cuota, p_modelo, paso, stake, real, resultado, beneficio, saldo)
  · Resumen     — KPIs: win rate, beneficio, ROI, drawdown, rachas…
  · Equity      — curva de saldo con gráfico de línea
  · TablaApuestas — la tabla de apuestas de referencia
  · Notas       — supuestos y limitaciones de la simulación

USO:
  python3 generar_excel_resultados.py --csv resultados_simulacion.csv
"""
import argparse
import csv
import math
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AZUL = "16213B"
AZUL2 = "0057FF"
AMAR = "FFF3C4"
VERDE = "E2F4E6"
ROJO = "FBE3E4"
NARANJA = "FDE9D9"
GRIS = "F0F2F7"
thin = Side(style="thin", color="C9CFDD")
borde = Border(left=thin, right=thin, top=thin, bottom=thin)


def est_cab(c):
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borde


def titulo(ws, txt, ncols):
    ws["A1"] = txt
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)


def generar(csv_ruta="resultados_simulacion.csv", bankroll=500.0,
            titulo_extra="", salida="Resultados_Simulacion.xlsx"):
    with open(csv_ruta, newline="", encoding="utf-8") as f:
        trades = list(csv.DictReader(f))

    wb = Workbook()

    # ====================================================== RESULTADOS
    ws = wb.active
    ws.title = "Resultados"
    titulo(ws, f"RESULTADOS DE LA SIMULACIÓN (paper trading)  ·  {titulo_extra}", 12)
    ws["A2"] = ("Saldo inicial: $%.2f · Sin comisiones · Precio de entrada fijo "
                "(cuota = 1/precio) · Resolución con el total real de la ventana de 48 h "
                "(suma de los 2 días siguientes del CSV)." % bankroll)
    ws["A2"].font = Font(size=9, italic=True, color="555555")
    ws.merge_cells("A2:L2")
    cab = ["#", "Fecha entrada", "Bin", "Lado", "Precio", "Cuota", "p_modelo",
           "Paso", "Stake ($)", "Real (48h)", "Resultado", "Beneficio ($)", "Saldo ($)"]
    for j, h in enumerate(cab, start=1):
        est_cab(ws.cell(row=4, column=j, value=h))
    for i, tr in enumerate(trades):
        r = 5 + i
        vals = [i + 1, tr["fecha"], tr["bin"], tr["lado"], float(tr["precio"]),
                float(tr["cuota"]), float(tr["p_modelo"]), int(tr["paso"]),
                float(tr["stake"]), int(tr["real"]), tr["resultado"],
                float(tr["beneficio"]), float(tr["saldo"])]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = borde
            if j in (5, 6, 9, 12, 13):
                c.number_format = "0.00"
            if j == 7:
                c.number_format = "0.0%"
            if j == 11:
                c.fill = PatternFill("solid", fgColor=VERDE if v == "G" else ROJO)
                c.font = Font(bold=True, color="1E7A34" if v == "G" else "C0392B")
    for col, w in zip("ABCDEFGHIJKLM", [5, 13, 9, 6, 8, 8, 9, 6, 9, 9, 10, 11, 11]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    n = len(trades)
    if n == 0:
        ws["A5"] = "Sin operaciones en el periodo simulado."
        ws["A5"].font = Font(italic=True, color="888888")

    # ====================================================== RESUMEN (KPIs)
    ws = wb.create_sheet("Resumen")
    titulo(ws, "PANEL DE RESULTADOS — estrategia Polymarket · @elonmusk", 2)
    ws["A3"] = "Bankroll inicial ($)"
    ws["B3"] = bankroll
    ws["B3"].fill = PatternFill("solid", fgColor=AMAR)
    ws["B3"].border = borde
    ws["B3"].number_format = "#,##0.00"
    fin_res = 4 + max(n, 1)  # última fila de Resultados (aunque esté vacía)
    filas_kpi = [
        ("Apuestas simuladas", f"=IF(Resultados!A5=\"\",0,COUNT(Resultados!A5:A{fin_res}))", "0"),
        ("Ganadas (G)", f'=COUNTIF(Resultados!K5:K{fin_res},"G")', "0"),
        ("Perdidas (P)", f'=COUNTIF(Resultados!K5:K{fin_res},"P")', "0"),
        ("WIN RATE", f"=IF(B6=0,\"—\",B7/B6)", "0.0%"),
        ("Beneficio neto ($)", f"=SUM(Resultados!L5:L{fin_res})", "#,##0.00"),
        ("ROI sobre bankroll", '=IF(B3>0,B10/B3,"—")', "0.0%"),
        ("Saldo final ($)", "=B3+B10", "#,##0.00"),
        ("Racha máx. de pérdidas", f"=MAX(Resultados!L5:L{fin_res}) - MIN(Resultados!L5:L{fin_res})", "—"),
        ("Cuota media", f'=IF(B6=0,"—",AVERAGEIF(Resultados!F5:F{fin_res},">0",Resultados!F5:F{fin_res}))', "0.00"),
        ("Total invertido ($)", f"=SUM(Resultados!I5:I{fin_res})", "#,##0.00"),
        ("Beneficio por apuesta ($)", '=IF(B6=0,"—",B10/B6)', "0.00"),
    ]
    for i, (label, formula, fmt) in enumerate(filas_kpi):
        r = 6 + i
        ws.cell(row=r, column=1, value=label).font = Font(size=10, bold=(i == 3))
        c = ws.cell(row=r, column=2, value=formula)
        c.border = borde
        if fmt:
            c.number_format = fmt
        if i == 3:
            c.font = Font(bold=True, size=12, color=AZUL2)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws["A19"] = ("ADVERTENCIA: los resultados son SIMULADOS (paper trading). "
                 "No garantizan resultados futuros. Muestra pequeña = baja significancia.")
    ws["A19"].font = Font(size=9, italic=True, color="B04A0A")

    # ====================================================== EQUITY (curva)
    ws = wb.create_sheet("Equity")
    titulo(ws, "CURVA DE SALDO (equity) — operación a operación", 3)
    ws["A3"], ws["B3"], ws["C3"] = "#", "Saldo ($)", "Beneficio acumulado ($)"
    for j in range(1, 4):
        est_cab(ws.cell(row=3, column=j))
    for i, tr in enumerate(trades):
        r = 4 + i
        ws.cell(row=r, column=1, value=i + 1).border = borde
        c = ws.cell(row=r, column=2, value=float(tr["saldo"]))
        c.border = borde
        c.number_format = "#,##0.00"
        c = ws.cell(row=r, column=3, value=float(tr["beneficio"]))
        c.border = borde
        c.number_format = "#,##0.00"
    if trades:
        chart = LineChart()
        chart.title = "Evolución del saldo (paper trading)"
        chart.style = 12
        chart.y_axis.title = "Saldo ($)"
        chart.x_axis.title = "Operación"
        chart.width, chart.height = 22, 12
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(trades))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(trades))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.solidFill = AZUL2
        chart.series[0].graphicalProperties.line.width = 22000
        ws.add_chart(chart, "E5")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20

    # ====================================================== TABLA APUESTAS
    ws = wb.create_sheet("TablaApuestas")
    titulo(ws, "TABLA DE APUESTAS DE REFERENCIA (progresión 1.5× desde $3.30)", 4)
    for j, h in enumerate(["Paso", "Stake ($)", "Pérdida acumulada ($)", "Neto si gana (cuota 3)"], start=1):
        est_cab(ws.cell(row=3, column=j, value=h))
    perd_prev = 0.0
    for i in range(1, 8):
        r = 3 + i
        stake = round(3.3 * 1.5 ** (i - 1), 2)
        perd = perd_prev + stake
        neto = round(stake * 2 - perd_prev, 2)
        vals = [i, stake, perd, neto]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = borde
            if j > 1:
                c.number_format = "#,##0.00"
        if i == 7:
            for j in range(1, 5):
                ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=NARANJA)
                ws.cell(row=r, column=j).font = Font(bold=True)
        perd_prev = perd
    for col, w in zip("ABCD", [8, 12, 22, 24]):
        ws.column_dimensions[col].width = w

    # ====================================================== NOTAS
    ws = wb.create_sheet("Notas")
    titulo(ws, "SUPUESTOS Y LIMITACIONES DE LA SIMULACIÓN", 1)
    notas = [
        "1. Mercado: se asume un mercado de 48 h disponible cada 2 días, cuyo bin [A,B] es el simulado.",
        "2. Precio de entrada: fijo e igual al --precio indicado (cuota = 1/precio). Sin comisiones, sin deslizamiento (salvo --slippage).",
        "3. Resolución: el total real de la ventana se calcula como la suma de los 2 días siguientes del CSV (días completos, hora ET).",
        "4. Lado: automático por defecto (YES si p_modelo≥60% y cuota≥3; NO si p_modelo≤30% y cuota NO≥3).",
        "5. Regla secuencial: no se abre una nueva apuesta hasta que la anterior se ha resuelto.",
        "6. Progresión: stake = $3.30 × 1.5^(paso−1); reinicio a $3.30 tras ganar; stop de ciclo en el paso 7 con pausa de 1 día.",
        "7. Los datos pueden ser reconstruidos (ver fuentes_serie.json) o parciales: los resultados dependen de su calidad.",
        "8. Esta simulación NO es asesoramiento financiero; el riesgo real incluye liquidez, slippage y errores de resolución.",
        "9. Para probar en vivo sin dinero: papel.py abre y resuelve apuestas simuladas con las señales reales.",
    ]
    for i, n in enumerate(notas):
        ws.cell(row=3 + i, column=1, value=n).font = Font(size=10)
    ws.column_dimensions["A"].width = 120

    wb.save(salida)
    return salida


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Genera Excel de resultados desde un CSV de operaciones")
    ap.add_argument("--csv", default="resultados_simulacion.csv")
    ap.add_argument("--bankroll", type=float, default=500.0)
    ap.add_argument("--titulo", default="")
    args = ap.parse_args()
    ruta = generar(args.csv, bankroll=args.bankroll, titulo_extra=args.titulo)
    print("Excel generado:", ruta)
