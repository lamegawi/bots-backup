#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXCEL HISTORIAL — libro que VA AÑADIENDO todas las operaciones
==============================================================
Mantiene Historial_Operaciones.xlsx como registro ACUMULATIVO:
cada ejecución lee el CSV de operaciones, compara con lo ya guardado
en el Excel (por ID) y SOLO AÑADE las operaciones nuevas. El Resumen
(KPIs) y la curva de Equity se recalculan siempre con TODAS las filas.

Fuentes:
  · resultados_papel.csv      → operaciones de PAPEL (el bot)
  · resultados_simulacion.csv → operaciones de SIMULACIÓN (simulador.py)
  · cualquier CSV con el mismo formato (columna "id" opcional)

USO:
  python3 excel_historial.py                          # añade las de papel
  python3 excel_historial.py --csv resultados_simulacion.csv
  python3 excel_historial.py --salida MiLibro.xlsx
"""
import argparse
import csv
import hashlib
import os
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    ET = ZoneInfo("America/New_York")
except Exception:
    # Windows sin tzdata: fallback a hora fija de verano (UTC-4)
    from datetime import timedelta
    from datetime import timezone as _tz
    ET = _tz(timedelta(hours=-4), name="EDT")
AZUL = "16213B"
AZUL2 = "0057FF"
AMAR = "FFF3C4"
VERDE = "E2F4E6"
ROJO = "FBE3E4"
NARANJA = "FDE9D9"
thin = Side(style="thin", color="C9CFDD")
borde = Border(left=thin, right=thin, top=thin, bottom=thin)
CAB = ["ID", "#", "Fecha", "Tipo", "Mercado", "Bin", "Lado", "Precio", "Cuota",
       "p_modelo", "Paso", "Stake ($)", "Real", "Resultado", "Beneficio ($)", "Saldo ($)"]


def est_cab(c):
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borde


def titulo(ws, txt, ncols):
    ws["A1"] = txt
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)


def leer_csv(csv_ruta, tipo_por_defecto):
    filas = []
    if not os.path.exists(csv_ruta):
        return filas
    with open(csv_ruta, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if not r.get("fecha"):
                continue
            filas.append(r)
    for r in filas:
        if not r.get("id"):
            clave = "|".join(str(r.get(k, "")) for k in
                             ("fecha", "bin", "lado", "precio", "paso", "stake",
                              "real", "resultado", "beneficio", "saldo"))
            r["id"] = hashlib.sha1(clave.encode()).hexdigest()[:12]
        r["_tipo"] = (r.get("tipo") or tipo_por_defecto).upper()
        r.setdefault("mercado", "—")
    return filas


def cargar_existentes(salida):
    """Lee los IDs ya guardados en el Excel si existe."""
    if not os.path.exists(salida):
        return set(), None
    try:
        wb = load_workbook(salida)
        if "Operaciones" not in wb.sheetnames:
            return set(), None
        ws = wb["Operaciones"]
        ids = set()
        for fila in ws.iter_rows(min_row=5, max_col=1, values_only=True):
            if fila[0]:
                ids.add(str(fila[0]))
        return ids, wb
    except Exception:
        return set(), None


def construir_libro(todas, bankroll, titulo_extra, salida):
    """Reconstruye el libro completo con todas las filas acumuladas."""
    wb = Workbook()

    # ------------------------------------------------------ OPERACIONES
    ws = wb.active
    ws.title = "Operaciones"
    titulo(ws, f"HISTORIAL ACUMULATIVO DE OPERACIONES  ·  {titulo_extra}", len(CAB))
    ws["A2"] = (f"Actualizado: {datetime.now(ET).strftime('%Y-%m-%d %H:%M %Z')} · "
                f"Total: {len(todas)} operaciones · Este libro SOLO añade filas "
                "(nunca borra: cada operación tiene un ID único).")
    ws["A2"].font = Font(size=9, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(CAB))
    for j, h in enumerate(CAB, start=1):
        est_cab(ws.cell(row=4, column=j, value=h))
    for i, r in enumerate(todas):
        fila = 5 + i
        vals = [r["id"], i + 1, r["fecha"], r["_tipo"], r["mercado"], r.get("bin", "—"),
                r.get("lado", "—"), r.get("precio", ""), r.get("cuota", ""),
                r.get("p_modelo", ""), r.get("paso", ""), r.get("stake", ""),
                r.get("real", ""), r.get("resultado", ""), r.get("beneficio", ""),
                r.get("saldo", "")]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=fila, column=j, value=v)
            c.border = borde
            if j in (8, 9, 12, 15, 16):
                try:
                    c.value = float(v) if v not in (None, "") else v
                    c.number_format = "0.00"
                except Exception:
                    pass
            if j == 10:
                try:
                    c.value = float(v) if v not in (None, "") else v
                    c.number_format = "0.0%"
                except Exception:
                    pass
            if j == 14:
                c.font = Font(bold=True, color="1E7A34" if v == "G" else "C0392B")
                c.fill = PatternFill("solid", fgColor=VERDE if v == "G" else ROJO)
    for col, w in zip("ABCDEFGHIJKLMNOP",
                      [14, 5, 12, 10, 34, 9, 6, 8, 8, 9, 6, 10, 10, 10, 11, 11]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ------------------------------------------------------ RESUMEN
    ws = wb.create_sheet("Resumen")
    titulo(ws, "PANEL DE RESULTADOS ACUMULADOS", 2)
    ws["A3"] = "Bankroll inicial ($)"
    ws["B3"] = bankroll
    ws["B3"].fill = PatternFill("solid", fgColor=AMAR)
    ws["B3"].border = borde
    ws["B3"].number_format = "#,##0.00"
    FIL = 5 + len(todas) if todas else 5
    kpis = [
        ("Operaciones totales", f"=COUNTIF(Operaciones!N5:N{FIL},\"G\")+COUNTIF(Operaciones!N5:N{FIL},\"P\")", "0"),
        ("Ganadas (G)", f'=COUNTIF(Operaciones!N5:N{FIL},"G")', "0"),
        ("Perdidas (P)", f'=COUNTIF(Operaciones!N5:N{FIL},"P")', "0"),
        ("WIN RATE", "=IF(B6=0,\"—\",B7/B6)", "0.0%"),
        ("Beneficio neto ($)", f"=SUM(Operaciones!O5:O{FIL})", "#,##0.00"),
        ("ROI sobre bankroll", '=IF(B3>0,B10/B3,"—")', "0.0%"),
        ("Saldo actual ($)", "=B3+B10", "#,##0.00"),
        ("Total invertido ($)", f"=SUM(Operaciones!L5:L{FIL})", "#,##0.00"),
        ("Cuota media", f'=IF(B6=0,"—",AVERAGEIF(Operaciones!I5:I{FIL},">0",Operaciones!I5:I{FIL}))', "0.00"),
        ("Beneficio por apuesta ($)", '=IF(B6=0,"—",B10/B6)', "0.00"),
    ]
    for i, (label, formula, fmt) in enumerate(kpis):
        r = 6 + i
        ws.cell(row=r, column=1, value=label).font = Font(size=10, bold=(i == 3))
        c = ws.cell(row=r, column=2, value=formula)
        c.border = borde
        if fmt:
            c.number_format = fmt
        if i == 3:
            c.font = Font(bold=True, size=12, color=AZUL2)
    # métricas estáticas (se recalculan al regenerar el libro)
    if todas:
        saldos = [float(x.get("saldo") or 0) for x in todas]
        racha = max_consecutivas([x.get("resultado") for x in todas])
        dd = max_drawdown(saldos)
        ws.cell(row=18, column=1, value="Racha máx. de pérdidas (estático)").font = Font(size=10)
        ws.cell(row=18, column=2, value=racha).border = borde
        ws.cell(row=19, column=1, value="Drawdown máx. $ (estático)").font = Font(size=10)
        ws.cell(row=19, column=2, value=round(dd, 2)).border = borde
        ws["B19"].number_format = "#,##0.00"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws["A21"] = ("Este Excel es ACUMULATIVO: el bot añade cada operación resuelta "
                 "sin borrar las anteriores. Los KPIs se recalculan con todas las filas.")
    ws["A21"].font = Font(size=9, italic=True, color="555555")

    # ------------------------------------------------------ EQUITY
    ws = wb.create_sheet("Equity")
    titulo(ws, "CURVA DE SALDO (equity) — todas las operaciones", 3)
    ws["A3"], ws["B3"], ws["C3"] = "#", "Saldo ($)", "Beneficio por op. ($)"
    for j in range(1, 4):
        est_cab(ws.cell(row=3, column=j))
    for i, r in enumerate(todas):
        fila = 4 + i
        ws.cell(row=fila, column=1, value=i + 1).border = borde
        c = ws.cell(row=fila, column=2, value=float(r.get("saldo") or 0))
        c.border = borde
        c.number_format = "#,##0.00"
        c = ws.cell(row=fila, column=3, value=float(r.get("beneficio") or 0))
        c.border = borde
        c.number_format = "#,##0.00"
    if todas:
        chart = LineChart()
        chart.title = "Evolución del saldo (acumulado)"
        chart.style = 12
        chart.y_axis.title = "Saldo ($)"
        chart.x_axis.title = "Operación"
        chart.width, chart.height = 24, 13
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(todas))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(todas))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.solidFill = AZUL2
        chart.series[0].graphicalProperties.line.width = 22000
        ws.add_chart(chart, "E5")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20

    # ------------------------------------------------------ NOTAS
    ws = wb.create_sheet("Notas")
    titulo(ws, "CÓMO FUNCIONA ESTE LIBRO", 1)
    notas = [
        "· Cada operación tiene un ID único; el libro SOLO AÑADE filas nuevas (nunca duplica ni borra).",
        "· El bot lo actualiza automáticamente cada vez que se resuelve una apuesta de papel.",
        "· Puedes añadir manualmente operaciones desde cualquier CSV con el mismo formato:",
        "      python3 excel_historial.py --csv mi_archivo.csv",
        "· Los KPIs del Resumen se recalculan con todas las filas cada vez que se regenera.",
        "· Racha máx. y drawdown son estáticos (se calculan al regenerar).",
        "· Resultados SIMULADOS (paper trading) — no garantizan resultados reales.",
    ]
    for i, n in enumerate(notas):
        ws.cell(row=3 + i, column=1, value=n).font = Font(size=10)
    ws.column_dimensions["A"].width = 110

    wb.save(salida)
    return salida


def max_consecutivas(resultados):
    racha = mejor = 0
    for r in resultados:
        racha = racha + 1 if r == "P" else 0
        mejor = max(mejor, racha)
    return mejor


def max_drawdown(saldos):
    pico = saldos[0] if saldos else 0
    dd = 0.0
    for s in saldos:
        pico = max(pico, s)
        dd = max(dd, pico - s)
    return dd


def generar(csv_ruta="resultados_papel.csv", salida="Historial_Operaciones.xlsx",
            bankroll=500.0, titulo_extra=""):
    tipo = "PAPEL" if "papel" in os.path.basename(csv_ruta) else \
           "SIMULACION" if "simulacion" in os.path.basename(csv_ruta) else "GENERAL"
    nuevas = leer_csv(csv_ruta, tipo)
    existentes_ids, _ = cargar_existentes(salida)
    a_anadir = [r for r in nuevas if r["id"] not in existentes_ids]
    todas = []
    if os.path.exists(salida):
        try:
            wb_prev = load_workbook(salida)
            ws_prev = wb_prev["Operaciones"]
            for fila in ws_prev.iter_rows(min_row=5, values_only=True):
                if not fila[0]:
                    continue
                todas.append({
                    "id": str(fila[0]), "fecha": fila[2] or "", "_tipo": fila[3] or "",
                    "mercado": fila[4] or "—", "bin": fila[5] or "—", "lado": fila[6] or "—",
                    "precio": fila[7], "cuota": fila[8], "p_modelo": fila[9],
                    "paso": fila[10], "stake": fila[11], "real": fila[12],
                    "resultado": fila[13], "beneficio": fila[14], "saldo": fila[15]})
        except Exception:
            todas = []
    todas += a_anadir
    construir_libro(todas, bankroll, titulo_extra, salida)
    return salida, len(a_anadir), len(todas)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Excel historial acumulativo de operaciones")
    ap.add_argument("--csv", default="resultados_papel.csv", help="CSV fuente de operaciones")
    ap.add_argument("--salida", default="Historial_Operaciones.xlsx")
    ap.add_argument("--bankroll", type=float, default=500.0)
    ap.add_argument("--titulo", default="")
    args = ap.parse_args()
    ruta, anadidas, total = generar(args.csv, salida=args.salida,
                                    bankroll=args.bankroll, titulo_extra=args.titulo)
    print(f"Excel: {ruta}")
    print(f"  · añadidas ahora: {anadidas} · total en el libro: {total}")
