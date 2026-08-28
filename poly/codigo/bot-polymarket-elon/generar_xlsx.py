#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera Tabla_Apuestas_Elon.xlsx — libro de control de la estrategia.
NOTA: las fórmulas se guardan en formato canónico inglés (SUM, IF, COUNTIF...);
Excel las muestra traducidas al español automáticamente."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

AZUL   = "16213B"
AZUL2  = "0057FF"
AMAR   = "FFF3C4"
GRIS   = "F0F2F7"
NARANJA= "FDE9D9"
VERDE  = "E2F4E6"
ROJO   = "FBE3E4"
BLANCO = "FFFFFF"

thin = Side(style="thin", color="C9CFDD")
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

def est_cab(c):
    c.font = Font(bold=True, color=BLANCO, size=10)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borde

def titulo(ws, txt, ncols):
    ws["A1"] = txt
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

def cabecera(ws, fila, encabezados):
    for j, h in enumerate(encabezados, start=1):
        est_cab(ws.cell(row=fila, column=j, value=h))

wb = Workbook()

# ============================================================ TABLA APUESTAS
ws = wb.active
ws.title = "TablaApuestas"
titulo(ws, "TABLA DE APUESTAS — Polymarket · Elon Musk # tweets (ventanas de 48 h)", 6)
ws["A2"] = ("Progresión: stake = $3.30 × 1.5^(paso−1) · reinicio a $3.30 tras ganar · "
            "cuota mínima 3.00 (precio ≤ 0.33) · stop-loss de ciclo en el paso 7 · "
            "una sola apuesta activa (la siguiente solo tras resolver la anterior).")
ws["A2"].font = Font(size=9, italic=True, color="555555")
ws.merge_cells("A2:F2")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 28

cabecera(ws, 4, ["Paso", "Stake ($)", "Pérdida acumulada si falla ($)",
                 "Beneficio neto si acierta — cuota 3.00 ($)",
                 "Beneficio neto si acierta — cuota 4.00 ($)", "Riesgo total del ciclo ($)"])
for i in range(1, 21):
    r = 4 + i
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=f"=ROUNDUP(3.3*1.5^(A{r}-1),2)")
    ws.cell(row=r, column=3, value=f"=SUM($B$5:B{r})")
    ws.cell(row=r, column=4, value=f"=ROUNDUP(B{r}*2-IF(A{r}=1,0,C{r-1}),2)")
    ws.cell(row=r, column=5, value=f"=ROUNDUP(B{r}*3-IF(A{r}=1,0,C{r-1}),2)")
    ws.cell(row=r, column=6, value=f"=C{r}")
    for j in range(1, 7):
        c = ws.cell(row=r, column=j)
        c.border = borde
        c.alignment = Alignment(horizontal="center" if j == 1 else "right")
        if j > 1:
            c.number_format = "#,##0.00"
    if i == 7:
        for j in range(1, 7):
            ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=NARANJA)
            ws.cell(row=r, column=j).font = Font(bold=True)
for col, w in zip("ABCDEF", [8, 12, 22, 30, 30, 22]):
    ws.column_dimensions[col].width = w
ws["A27"] = ("PASO 7 = LÍMITE (stop-loss). Si falla el paso 7, el ciclo se cierra asumiendo la pérdida "
             "acumulada de $106.18 y se reinicia en $3.30 tras una pausa de 24 h.")
ws["A27"].font = Font(bold=True, color="B04A0A", size=10)
ws["A28"] = ("Punto de equilibrio: p* = 33,3% por apuesta (1 / cuota 3.00). La regla de entrada exige "
             "p_modelo ≥ 60% (YES) o ≤ 30% (NO), por lo que el margen de seguridad es amplio.")
ws["A28"].font = Font(size=9, italic=True, color="555555")

# ============================================================ SEÑAL
ws = wb.create_sheet("Senal")
titulo(ws, "CALCULADOR DE SEÑAL — introduce los datos amarillos", 3)
ws["A2"] = ("Fuente de datos: Social Blade (gráfico Tweets Posted Weekly) o conteo diario de @elonmusk "
            "a las 23:59 ET. Conteo = posts + quote posts + reposts (mismo criterio que el mercado).")
ws["A2"].font = Font(size=9, italic=True, color="555555")

entradas = [
    ("AVG7 — media de tweets/día de los últimos 7 días completos", 12.5),
    ("V2 — total de tweets de los últimos 2 días completos", 35),
    ("Tweets ya publicados dentro de la ventana del mercado (T0)", 0),
    ("Horas transcurridas de la ventana de 48 h (0–48)", 0),
    ("A — límite inferior del bin del mercado", 50),
    ("B — límite superior del bin (9999 = '≥ A')", 9999),
    ("Precio actual del YES (0–1)", 0.33),
    ("Paso actual del ciclo (1–7)", 1),
]
for i, (label, val) in enumerate(entradas):
    r = 4 + i
    ws.cell(row=r, column=1, value=label).font = Font(size=10)
    c = ws.cell(row=r, column=2, value=val)
    c.fill = PatternFill("solid", fgColor=AMAR)
    c.border = borde
    c.number_format = "0.00" if i in (0, 6, 7) else "0"

calc = [
    ("R = V2 / (2·AVG7)", '=IF(B4>0,B5/(2*B4),"n/d")', "0.00"),
    ("ajuste(R) = MIN(1.5; MAX(0.5; 1+0.5·(R−1)))", "=MIN(1.5,MAX(0.5,1+0.5*(B12-1)))", "0.000"),
    ("λ48 = 2 · AVG7 · ajuste  (tweets esperados en 48 h)", "=2*B4*B13", "0.0"),
    ("λ restante = λ48 · (48−horas)/48", "=B14*MAX(0,(48-B7))/48", "0.0"),
    ("p_modelo = P(A ≤ X ≤ B)  con X ~ Poisson(λ restante)", "=IF(B15>0,POISSON.DIST(B9-B6,B15,TRUE)-POISSON.DIST(B8-1-B6,B15,TRUE),0)", "0.0%"),
    ("Cuota YES = 1 / precio", '=IF(B10>0,1/B10,"—")', "0.00"),
    ("Precio NO = 1 − precio YES", "=1-B10", "0.000"),
    ("Cuota NO = 1 / precio NO", '=IF(B18>0,1/B18,"—")', "0.00"),
    ("Cuota mínima exigida", '=IF(AND(B10<=0.333,B18<=0.333),"ambos lados OK",IF(B10<=0.333,"solo YES",IF(B18<=0.333,"solo NO","NINGÚN lado (PASAR)")))', ""),
    ("► VEREDICTO", '=IF(B4<5,"PASAR — AVG7 < 5 (base insuficiente)",IF(AND(B16>=0.6,B10<=0.333),"APOSTAR YES",IF(AND(B16<=0.3,B18<=0.333),"APOSTAR NO","PASAR — sin ventaja")))', ""),
    ("Stake sugerido ($) según el paso del ciclo", '=IF(LEFT(B21,6)="APOSTA",VLOOKUP(B11,TablaApuestas!$A$5:$B$24,2,FALSE),"—")', "0.00"),
]
for i, (label, formula, fmt) in enumerate(calc):
    r = 12 + i
    ws.cell(row=r, column=1, value=label).font = Font(size=10)
    c = ws.cell(row=r, column=2, value=formula)
    c.border = borde
    if fmt:
        c.number_format = fmt
    if "VEREDICTO" in label:
        c.font = Font(bold=True, size=12, color=AZUL2)
ws.column_dimensions["A"].width = 62
ws.column_dimensions["B"].width = 26
ws["A26"] = ("Criterios: (1) apostar YES solo si p_modelo ≥ 60% y precio YES ≤ 0.333; "
             "(2) apostar NO solo si p_modelo ≤ 30% y precio NO ≤ 0.333; (3) en cualquier otro caso PASAR. "
             "Recuerda: una sola apuesta activa y espera a que el mercado se resuelva.")
ws["A26"].font = Font(size=9, italic=True, color="555555")

# ============================================================ REGISTRO
ws = wb.create_sheet("Registro")
titulo(ws, "REGISTRO DE APUESTAS — anota cada operación el mismo día (auditable)", 18)
cabecera(ws, 3, ["Fecha entrada", "Mercado (URL)", "Ventana", "Lado", "Precio", "Cuota",
                 "AVG7", "V2", "R", "p_modelo", "Señal OK", "Paso", "Stake ($)",
                 "Resultado", "Beneficio ($)", "Saldo ($)", "Notas", "RachaP (aux)"])
for r in range(4, 104):
    ws.cell(row=r, column=6, value=f'=IF(E{r}="","",ROUND(1/E{r},2))')
    ws.cell(row=r, column=13, value=f'=IF(L{r}="","",VLOOKUP(L{r},TablaApuestas!$A$5:$B$24,2,FALSE))')
    ws.cell(row=r, column=15, value=f'=IF(N{r}="G",M{r}*(F{r}-1),IF(N{r}="P",-M{r},0))')
    ws.cell(row=r, column=16, value=f'=IF(O{r}="","",KPIs!$B$3+SUM($O$4:O{r}))')
    ws.cell(row=r, column=18, value=f'=IF(N{r}="P",IF(ROW()=4,1,R{r-1}+1),0)')
    for j in range(1, 19):
        c = ws.cell(row=r, column=j)
        c.border = borde
    for j in (5, 7, 8, 9, 10, 13, 15, 16):
        ws.cell(row=r, column=j).number_format = "0.00"
for col, w in zip("ABCDEFGHIJKLMNOPQ", [12, 34, 16, 8, 8, 8, 8, 8, 8, 9, 9, 7, 10, 12, 12, 12, 28]):
    ws.column_dimensions[col].width = w
ws.column_dimensions["R"].width = 7
dv = DataValidation(type="list", formula1='"G,P,Pendiente"', allow_blank=True)
ws.add_data_validation(dv)
dv.add("N4:N103")
ws.conditional_formatting.add("N4:N103", CellIsRule(operator="equal", formula=['"G"'], fill=PatternFill("solid", fgColor=VERDE)))
ws.conditional_formatting.add("N4:N103", CellIsRule(operator="equal", formula=['"P"'], fill=PatternFill("solid", fgColor=ROJO)))
ws.conditional_formatting.add("O4:O103", CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="1E7A34", bold=True)))
ws.conditional_formatting.add("O4:O103", CellIsRule(operator="lessThan", formula=["0"], font=Font(color="C0392B", bold=True)))
ws.freeze_panes = "A4"
ej = [
    # A=fecha, B=mercado, C=ventana, D=lado, E=precio, G=AVG7, H=V2, I=R,
    # J=p_modelo, K=señal OK, L=paso, N=resultado, Q=notas  (F/M/O/P/R son fórmulas)
    {"A": "2026-08-05", "B": "polymarket.com/event/…", "C": "48 h", "D": "YES", "E": 0.33,
     "G": 30.9, "H": 95, "I": 1.54, "J": 1.00, "K": "SÍ", "L": 2, "N": "G", "Q": "EJEMPLO — borrar"},
    {"A": "2026-08-08", "B": "polymarket.com/event/…", "C": "48 h", "D": "YES", "E": 0.31,
     "G": 30.9, "H": 95, "I": 1.54, "J": 1.00, "K": "SÍ", "L": 1, "N": "Pendiente", "Q": "EJEMPLO — borrar"},
]
for i, fila in enumerate(ej):
    r = 4 + i
    for col, v in fila.items():
        ws[f"{col}{r}"] = v
        ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=GRIS)
ws["A6"] = "Introduce aquí tus apuestas reales (puedes borrar las filas de ejemplo)."

# ============================================================ KPIs
ws = wb.create_sheet("KPIs")
titulo(ws, "PANEL DE MEDICIÓN — actualiza el banco inicial", 2)
ws["A3"] = "Bankroll inicial ($)"
ws["B3"] = 500
c = ws["B3"]
c.fill = PatternFill("solid", fgColor=AMAR)
c.border = borde
c.number_format = "#,##0.00"
kpis = [
    ("Apuestas resueltas", '=COUNTIF(Registro!N4:N103,"G")+COUNTIF(Registro!N4:N103,"P")', "0"),
    ("Ganadas (G)", '=COUNTIF(Registro!N4:N103,"G")', "0"),
    ("Perdidas (P)", '=COUNTIF(Registro!N4:N103,"P")', "0"),
    ("Pendientes", '=COUNTIF(Registro!N4:N103,"Pendiente")', "0"),
    ("WIN RATE", '=IF(B6=0,"—",B7/B6)', "0.0%"),
    ("Beneficio neto ($)", "=SUM(Registro!O4:O103)", "#,##0.00"),
    ("ROI sobre bankroll", '=IF(B3>0,B11/B3,"—")', "0.0%"),
    ("Saldo actual ($)", "=B3+B11", "#,##0.00"),
    ("Racha máx. de pérdidas", "=MAX(Registro!R4:R103)", "0"),
    ("Peor saldo vs. inicial ($)", "=MIN(Registro!P4:P103)-B3", "#,##0.00"),
    ("Ciclos iniciados", "=COUNTIF(Registro!L4:L103,1)", "0"),
    ("Cuota media (resueltas)", '=IF(B6=0,"—",SUMPRODUCT((Registro!N4:N103="G")+(Registro!N4:N103="P"),Registro!F4:F103)/B6)', "0.00"),
    ("Total invertido ($)", "=SUM(Registro!M4:M103)", "#,##0.00"),
]
for i, (label, formula, fmt) in enumerate(kpis):
    r = 6 + i
    ws.cell(row=r, column=1, value=label).font = Font(size=10, bold=(i == 4))
    c = ws.cell(row=r, column=2, value=formula)
    c.border = borde
    if fmt:
        c.number_format = fmt
    if i == 4:
        c.font = Font(bold=True, size=12, color=AZUL2)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 16
ws["A21"] = ("Reglas de revisión: si tras 30 días el win rate < 40% o el beneficio neto < 0, "
             "PAUSA la estrategia y recalibra los umbrales (p_modelo, bins).")
ws["A21"].font = Font(size=9, italic=True, color="B04A0A")
ws["A22"] = ("Meta del diseño: win rate ≥ 60% (por el filtro de entrada p_modelo ≥ 0.60) "
             "y beneficio esperado ≈ +$6.60 por ciclo ganado (cuota 3.00).")
ws["A22"].font = Font(size=9, italic=True, color="555555")

# ============================================================ REGLAS
ws = wb.create_sheet("Reglas")
titulo(ws, "REGLAS DE LA ESTRATEGIA — checklist obligatorio antes de cada apuesta", 3)
cabecera(ws, 3, ["#", "Regla", "Cumplido (SÍ/NO)"])
reglas = [
    "Solo mercados Polymarket 'Elon Musk # tweets' de ventana 48 h con reglas de resolución claras (Social Blade / X, conteo = posts + quote posts + reposts).",
    "Volumen del mercado ≥ $5.000 y liquidez ≥ $1.000.",
    "Cuota del lado elegido ≥ 3.00 (precio ≤ 0.333). Si no se cumple, PASAR.",
    "Señal: p_modelo ≥ 60% para YES, o p_modelo ≤ 30% para NO (calculado con senal.py o la hoja Senal).",
    "AVG7 ≥ 5 tweets/día (base de datos mínima).",
    "UNA sola apuesta activa: no se abre la siguiente hasta que la anterior se ha RESUELTO.",
    "Progresión: 3.30 → ×1.50 tras cada fallo; tras ganar, reiniciar en 3.30.",
    "Stop-loss de ciclo: paso 7 como máximo; tras 7 fallos, cerrar ciclo y pausar 24 h.",
    "Nunca perseguir precios: si la cuota baja de 3.00 tras la señal, PASAR.",
    "Entrar solo en la primera mitad de la ventana (horas transcurridas ≤ 24).",
    "Bankroll mínimo $500; el riesgo máximo de un ciclo completo ($106.18) ≤ 25% del bankroll.",
    "Registrar cada apuesta en la hoja Registro el mismo día (fecha, precio, señal, paso, resultado).",
    "Revisión mensual de KPIs: si win rate < 40% o beneficio < 0 en 30 días → pausar y recalibrar.",
    "Operar solo con dinero que puedas permitirte perder por completo.",
]
for i, rtxt in enumerate(reglas):
    r = 4 + i
    ws.cell(row=r, column=1, value=i + 1)
    ws.cell(row=r, column=2, value=rtxt).alignment = Alignment(wrap_text=True, vertical="top")
    for j in range(1, 4):
        ws.cell(row=r, column=j).border = borde
    ws.row_dimensions[r].height = 28
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 105
ws.column_dimensions["C"].width = 16

wb.save("Tabla_Apuestas_Elon.xlsx")
print("OK — Tabla_Apuestas_Elon.xlsx generado")
